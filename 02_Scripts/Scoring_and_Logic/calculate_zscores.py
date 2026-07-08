import openpyxl
import numpy as np
import os

xlsx_path = "/home/dilshan/Desktop/Thallasemia research/Thalassemia_Research.xlsx"

print("Loading workbook...")
wb = openpyxl.load_workbook(xlsx_path, read_only=False)
sheet = wb.worksheets[0]

print("Reading rows...")
# Read all values sequentially to avoid slow random cell access
rows = list(sheet.iter_rows(values_only=True))
headers = rows[0]

q16_indices = []
q27_indices = []

for idx, h in enumerate(headers):
    if not h:
        continue
    # Match Q16 columns (excluding "I don’t know")
    if '16. What are the clinical forms' in h and 'I don’t know' not in h:
        q16_indices.append(idx)
    # Match Q27 columns
    if '27. Problems faced by thalassemia major' in h:
        q27_indices.append(idx)

print("Matched Q16 Columns indices:", q16_indices)
print("Matched Q27 Columns indices:", q27_indices)

# Calculate Knowledge Score for each row
scores = []
for r_idx in range(1, len(rows)):
    row_data = rows[r_idx]
    row_score = 0
    
    # Sum Q16 checked options
    for idx in q16_indices:
        val = row_data[idx]
        try:
            row_score += float(val) if val is not None else 0
        except (ValueError, TypeError):
            pass
            
    # Sum Q27 checked options
    for idx in q27_indices:
        val = row_data[idx]
        try:
            row_score += float(val) if val is not None else 0
        except (ValueError, TypeError):
            pass
            
    scores.append(row_score)

mean_score = np.mean(scores)
std_score = np.std(scores, ddof=1) # sample standard deviation

print(f"\nTotal participants scored: {len(scores)}")
print(f"Mean Knowledge Score: {mean_score:.4f}")
print(f"Standard Deviation: {std_score:.4f}")
print(f"Min Score: {min(scores)} | Max Score: {max(scores)}")

# Write headers and values back to the sheet
# We add columns at max_col + 1 and max_col + 2
max_col = len(headers)
score_col_idx = max_col + 1
zscore_col_idx = max_col + 2

# Write headers
sheet.cell(row=1, column=score_col_idx, value="Knowledge_Score")
sheet.cell(row=1, column=zscore_col_idx, value="Knowledge_Score_Z_Score")

# Write values
z_scores = []
for i, score in enumerate(scores):
    r_idx = i + 2 # row 2 is index 0 in scores list
    
    # Calculate Z-score
    if std_score > 0:
        z = (score - mean_score) / std_score
    else:
        z = 0.0
    z_scores.append(z)
    
    sheet.cell(row=r_idx, column=score_col_idx, value=score)
    sheet.cell(row=r_idx, column=zscore_col_idx, value=z)

print(f"\nVerification metrics:")
print(f"  Mean of Z-scores: {np.mean(z_scores):.6f} (should be approx 0)")
print(f"  SD of Z-scores: {np.std(z_scores, ddof=1):.6f} (should be approx 1)")

print("\nSaving workbook...")
wb.save(xlsx_path)
print("Workbook saved successfully!")
